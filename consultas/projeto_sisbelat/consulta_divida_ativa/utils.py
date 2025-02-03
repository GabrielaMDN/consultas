import requests
from fpdf import FPDF
from openpyxl import Workbook
from tkinter import filedialog, messagebox
import logging
import os
from dotenv import load_dotenv

# Configurar logs
logging.basicConfig(level=logging.INFO)

# Carregar variáveis de ambiente
load_dotenv()

# Chave de autorização carregada do .env
AUTH_KEY = os.getenv("AUTH_KEY")


def obter_token():
    """
    Obtém o token de acesso do serviço Serpro.
    """
    url = "https://gateway.apiserpro.serpro.gov.br/token"
    headers = {
        "Authorization": f"Basic {AUTH_KEY}",
        "Content-Type": "application/x-www-form-urlencoded"
    }
    data = {
        "grant_type": "client_credentials"
    }
    
    try:
        response = requests.post(url, headers=headers, data=data)
        response.raise_for_status()
        response_json = response.json()

        # Validar se o token foi retornado
        if "access_token" in response_json:
            logging.info("Token obtido com sucesso.")
            return response_json["access_token"]
        else:
            logging.error("Token não encontrado na resposta.")
            messagebox.showerror("Erro ao obter token", "Token não encontrado na resposta da API.")
            return None
    except requests.exceptions.RequestException as e:
        logging.error(f"Erro ao obter token: {e}")
        messagebox.showerror("Erro ao obter token", str(e))
        return None


def consultar_api(token, cnpj):
    """
    Consulta a API usando o token obtido e retorna os dados do CNPJ.
    """
    url = f"https://gateway.apiserpro.serpro.gov.br/consulta-divida-ativa-df/api/v1/devedor/{cnpj}"
    headers = {
        "Authorization": f"Bearer {token}"
    }
    
    try:
        response = requests.get(url, headers=headers)
        response.raise_for_status()

        dados = response.json()
        if dados:
            logging.info(f"Dados retornados para o CNPJ {cnpj}.")
            return dados
        else:
            logging.warning(f"Nenhum dado encontrado para o CNPJ {cnpj}.")
            messagebox.showinfo("Nenhum dado encontrado", f"Nenhum dado encontrado para o CNPJ {cnpj}.")
            return None
    except requests.exceptions.HTTPError as e:
        if e.response.status_code == 404:
            logging.warning(f"CNPJ {cnpj} não encontrado na base da API.")
            messagebox.showinfo("CNPJ não encontrado", f"O CNPJ {cnpj} não foi encontrado na base da API.")
        else:
            logging.error(f"Erro na consulta da API: {e}")
            messagebox.showerror("Erro na consulta", str(e))
        return None

def organizar_resultados(cnpj, dados):
    """
    Organiza os resultados em um formato legível.
    """
    resultado = [f"Resultado para o CNPJ {cnpj}:\n"]
    if isinstance(dados, list):
        for index, item in enumerate(dados, start=1):
            resultado.append(f"Item {index}:")
            if isinstance(item, dict):
                for chave, valor in item.items():
                    resultado.append(f"  {chave}: {valor}")
            else:
                resultado.append(f"  {item}")
    elif isinstance(dados, dict):
        for chave, valor in dados.items():
            resultado.append(f"{chave}: {valor}")
    else:
        resultado.append("Formato inesperado de resposta da API.")
    return "\n".join(resultado)


def exportar_para_pdf(cnpj, dados):
    """
    Exporta os resultados para PDF.
    """
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)
    pdf.cell(200, 10, txt=f"Resultado para o CNPJ {cnpj}", ln=True, align='C')
    pdf.ln(10)

    resultado_formatado = organizar_resultados(cnpj, dados)
    for linha in resultado_formatado.split("\n"):
        pdf.cell(0, 10, txt=linha, ln=True)

    file_path = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF files", "*.pdf")])
    if file_path:
        pdf.output(file_path)
        logging.info(f"PDF salvo em {file_path}.")

def exportar_para_xls(cnpj, dados):
    """
    Exporta os resultados para XLS no formato de duas colunas: Campo e Valor.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"CNPJ {cnpj}"

    # Adicionar cabeçalho
    ws.append(["Campo", "Valor"])

    # Processar os dados
    def processar_dados(dados, prefixo=""):
        """
        Processa os dados para adicionar no formato 'Campo' e 'Valor'.
        """
        if isinstance(dados, dict):
            for chave, valor in dados.items():
                nome_campo = f"{prefixo}{chave}" if prefixo else chave
                if isinstance(valor, (dict, list)):
                    processar_dados(valor, prefixo=f"{nome_campo} -> ")
                else:
                    ws.append([nome_campo, valor])
        elif isinstance(dados, list):
            for index, item in enumerate(dados, start=1):
                nome_campo = f"{prefixo}Item {index}"
                if isinstance(item, (dict, list)):
                    processar_dados(item, prefixo=f"{nome_campo} -> ")
                else:
                    ws.append([nome_campo, item])

    # Processar os resultados
    processar_dados(dados)

    # Salvar o arquivo XLS
    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        wb.save(file_path)
        logging.info(f"Arquivo XLS salvo em {file_path}.")
