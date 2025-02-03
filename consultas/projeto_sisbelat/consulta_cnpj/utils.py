import logging
import requests
from fpdf import FPDF
from openpyxl import Workbook
from dotenv import load_dotenv
from tkinter import filedialog, messagebox
import os

# Configuração de logging para depuração
logging.basicConfig(level=logging.INFO)

# Dicionários com constantes
TIPOS_ESTABELECIMENTO = {
    "1": "Matriz",
    "2": "Filial"
}

PORTE_EMPRESA = {
    "01": "Microempresa - ME",
    "03": "Empresa de pequeno porte - EPP",
    "05": "Demais empresas"
}

TIPOS_SITUACAO_CADASTRAL = {
    "1": "Nula",
    "2": "Ativa",
    "3": "Suspensa",
    "4": "Inapta",
    "5": "Ativa Não Regular",
    "8": "Baixada"
}

TIPOS_SITUACAO_ESPECIAL = [
    "Início de Concordata",
    "Término de Concordata",
    "Em Liquidação",
    "Em Liquidação Extra-Judicial",
    "Falido",
    "Intervenção",
    "Financeiro e de Capitais",
    "Liquidação Judicial",
    "Liquidação Extra-Judicial",
    "Recuperação Judicial"
]

TIPOS_SOCIO = {
    "1": "Pessoa Jurídica",
    "2": "Pessoa Física",
    "3": "Sócio Estrangeiro"
}

TIPOS_QUALIFICACAO_SOCIO = {
    "05": "Administrador",
    "08": "Conselheiro de Administração",
    "10": "Diretor",
    "16": "Presidente",
    "17": "Procurador",
    "18": "Secretário",
    "20": "Sociedade Consorciada",
    "21": "Sociedade Filiada",
    "22": "Sócio",
    "23": "Sócio Capitalista",
    "24": "Sócio Comanditado",
    "25": "Sócio Comanditário",
    "26": "Sócio de Indústria",
    "28": "Sócio-Gerente",
    "29": "Sócio Incapaz ou Relat.Incapaz (exceto menor)",
    "30": "Sócio Menor (Assistido/Representado)",
    "31": "Sócio Ostensivo",
    "33": "Tesoureiro",
    "37": "Sócio Pessoa Jurídica Domiciliado no Exterior",
    "38": "Sócio Pessoa Física Residente ou Domiciliado no Exterior",
    "47": "Sócio Pessoa Física Residente no Brasil",
    "48": "Sócio Pessoa Jurídica Domiciliado no Brasil",
    "49": "Sócio-Administrador",
    "52": "Sócio com Capital",
    "53": "Sócio sem Capital",
    "54": "Fundador",
    "55": "Sócio Comanditado Residente no Exterior",
    "56": "Sócio Comanditário Pessoa Física Residente no Exterior",
    "57": "Sócio Comanditário Pessoa Jurídica Domiciliado no Exterior",
    "58": "Sócio Comanditário Incapaz",
    "59": "Produtor Rural",
    "63": "Cotas em Tesouraria",
    "65": "Titular Pessoa Física Residente ou Domiciliado no Brasil",
    "66": "Titular Pessoa Física Residente ou Domiciliado no Exterior",
    "67": "Titular Pessoa Física Incapaz ou Relativamente Incapaz (exceto menor)",
    "68": "Titular Pessoa Física Menor (Assistido/Representado)",
    "70": "Administrador Residente ou Domiciliado no Exterior",
    "71": "Conselheiro de Administração Residente ou Domiciliado no Exterior",
    "72": "Diretor Residente ou Domiciliado no Exterior",
    "73": "Presidente Residente ou Domiciliado no Exterior",
    "74": "Sócio-Administrador Residente ou Domiciliado no Exterior",
    "75": "Fundador Residente ou Domiciliado no Exterior",
    "76": "Protetor",
    "77": "Vice-Presidente",
    "78": "Titular Pessoa Jurídica Domiciliada no Brasil",
    "79": "Titular Pessoa Jurídica Domiciliada no Exterior"
}

TIPOS_QUALIFICACAO_REPRESENTANTE_LEGAL = {
    "05": "Administrador",
    "09": "Curador",
    "14": "Mãe",
    "15": "Pai",
    "17": "Procurador",
    "35": "Tutor"
}

# Carregar variáveis de ambiente
load_dotenv()
AUTH_KEY = os.getenv("AUTH_KEY")
# Função para obter o token
def obter_token():
    """
    Obtém o token de acesso do serviço Serpro.
    """
    url = "https://gateway.apiserpro.serpro.gov.br/token"
    headers = {
        "Authorization": f"Basic {AUTH_KEY}",
        "Content-Type": "application/x-www-form-urlencoded"
    }
    data = {
        "grant_type": "client_credentials"
    }

    try:
        response = requests.post(url, headers=headers, data=data)
        response.raise_for_status()
        return response.json().get("access_token")
    except requests.exceptions.RequestException as e:
        logging.error(f"Erro ao obter token: {e}")
        return None


# Função para consultar a API
def consultar_api(token, cnpj):
    url = f"https://gateway.apiserpro.serpro.gov.br/consulta-cnpj-df/v2/qsa/{cnpj}"
    headers = {"Authorization": f"Bearer {token}"}

    try:
        response = requests.get(url, headers=headers)
        response.raise_for_status()
        return response.json()
    except requests.RequestException as e:
        logging.error(f"Erro ao consultar API: {e}")
        return None


# Função para exportar os resultados para PDF
def exportar_para_pdf(cnpj, dados, file_path):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)
    pdf.cell(200, 10, txt=f"Resultado para o CNPJ {cnpj}", ln=True, align='C')
    pdf.ln(10)

    for chave, valor in dados.items():
        pdf.cell(0, 10, txt=f"{chave}: {valor}", ln=True)

    pdf.output(file_path)
    logging.info(f"PDF salvo em {file_path}")


def exportar_para_xls(cnpj, dados, file_path):
    """
    Exporta os resultados para um arquivo Excel.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"Resultado CNPJ {cnpj}"

    # Adiciona cabeçalho e dados
    ws.append(["Chave", "Valor"])
    for chave, valor in dados.items():
        # Lida com valores aninhados e traduzidos
        if isinstance(valor, dict):
            valor = "; ".join(f"{sub_chave}: {sub_valor}" for sub_chave, sub_valor in valor.items())
        elif isinstance(valor, list):
            valor = "; ".join(str(item) for item in valor)
        ws.append([chave, str(valor)])

    # Salvar o arquivo Excel
    wb.save(file_path)
    logging.info(f"Excel salvo em {file_path}")

def traduzir_dados(dados):
    """
    Traduz os dados retornados pela API substituindo códigos por textos descritivos.
    """
    # Caso a API não retorne dados
    if not dados:
        return "NADA CONSTA"

    if isinstance(dados, list):
        for item in dados:
            if isinstance(item, dict):
                # Traduzir 'tipoEstabelecimento'
                item['tipoEstabelecimento'] = TIPOS_ESTABELECIMENTO.get(
                    item.get('tipoEstabelecimento', ''), item.get('tipoEstabelecimento', 'Desconhecido')
                )

                # Traduzir 'porte'
                porte_empresa = item.get('porte', '')
                if isinstance(porte_empresa, str):
                    item['porte'] = PORTE_EMPRESA.get(porte_empresa, 'Desconhecido')

                # Traduzir 'situacaoCadastral'
                situacao_cadastral = item.get('situacaoCadastral', {})
                if isinstance(situacao_cadastral, dict):
                    codigo_situacao = situacao_cadastral.get('codigo', '')
                    situacao_cadastral['codigo'] = TIPOS_SITUACAO_CADASTRAL.get(
                        str(codigo_situacao), 'Desconhecido'
                    )
                    item['situacaoCadastral'] = situacao_cadastral

                # Traduzir lista de sócios
                socios = item.get('socios', [])
                if isinstance(socios, list):
                    for socio in socios:
                        # Traduzir 'tipoSocio'
                        print(f"Antes da tradução - tipoSocio: {socio.get('tipoSocio', '')}")
                        socio['tipoSocio'] = TIPOS_SOCIO.get(
                            str(socio.get('tipoSocio', '')), 'Desconhecido'
                        )
                        print(f"Depois da tradução - tipoSocio: {socio.get('tipoSocio', '')}")

                        # Traduzir 'qualificacao'
                        print(f"Antes da tradução - qualificacao: {socio.get('qualificacao', '')}")
                        socio['qualificacao'] = TIPOS_QUALIFICACAO_SOCIO.get(
                            str(socio.get('qualificacao', '')), 'Desconhecido'
                        )
                        print(f"Depois da tradução - qualificacao: {socio.get('qualificacao', '')}")

                        # Traduzir 'representanteLegal'
                        representante = socio.get('representanteLegal', {})
                        if isinstance(representante, dict):
                            print(f"Antes da tradução - representanteLegal.qualificacao: {representante.get('qualificacao', '')}")
                            representante['qualificacao'] = TIPOS_QUALIFICACAO_REPRESENTANTE_LEGAL.get(
                                str(representante.get('qualificacao', '')), 'Desconhecido'
                            )
                            print(f"Depois da tradução - representanteLegal.qualificacao: {representante.get('qualificacao', '')}")
                            socio['representanteLegal'] = representante

    elif isinstance(dados, dict):
        for chave, valor in dados.items():
            if chave == 'tipoEstabelecimento':
                dados[chave] = TIPOS_ESTABELECIMENTO.get(valor, 'Desconhecido')
            elif chave == 'porte':
                dados[chave] = PORTE_EMPRESA.get(str(valor), 'Desconhecido')
            elif chave == 'situacaoCadastral' and isinstance(valor, dict):
                codigo_situacao = valor.get('codigo', '')
                valor['codigo'] = TIPOS_SITUACAO_CADASTRAL.get(str(codigo_situacao), 'Desconhecido')
                dados[chave] = valor
            elif chave == 'socios' and isinstance(valor, list):
                for socio in valor:
                    # Traduzir 'tipoSocio'
                    print(f"Antes da tradução - tipoSocio: {socio.get('tipoSocio', '')}")
                    socio['tipoSocio'] = TIPOS_SOCIO.get(
                        str(socio.get('tipoSocio', '')), 'Desconhecido'
                    )
                    print(f"Depois da tradução - tipoSocio: {socio.get('tipoSocio', '')}")

                    # Traduzir 'qualificacao'
                    print(f"Antes da tradução - qualificacao: {socio.get('qualificacao', '')}")
                    socio['qualificacao'] = TIPOS_QUALIFICACAO_SOCIO.get(
                        str(socio.get('qualificacao', '')), 'Desconhecido'
                    )
                    print(f"Depois da tradução - qualificacao: {socio.get('qualificacao', '')}")

                    # Traduzir 'representanteLegal'
                    representante = socio.get('representanteLegal', {})
                    if isinstance(representante, dict):
                        print(f"Antes da tradução - representanteLegal.qualificacao: {representante.get('qualificacao', '')}")
                        representante['qualificacao'] = TIPOS_QUALIFICACAO_REPRESENTANTE_LEGAL.get(
                            str(representante.get('qualificacao', '')), 'Desconhecido'
                        )
                        print(f"Depois da tradução - representanteLegal.qualificacao: {representante.get('qualificacao', '')}")
                        socio['representanteLegal'] = representante

    return dados


