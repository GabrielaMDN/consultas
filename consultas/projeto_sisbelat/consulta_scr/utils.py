import requests
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import logging
from dotenv import load_dotenv
import os

# Carregar variáveis de ambiente
load_dotenv()
BASE_URL = os.getenv("BASE_URL")

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(message)s")

def obter_token(email, password):
    """
    Autentica o usuário e retorna o token de acesso.
    """
    login_url = f'{BASE_URL}/login'
    login_data = {'email': email, 'password': password}
    headers = {'Content-Type': 'application/x-www-form-urlencoded'}

    response = requests.post(login_url, data=login_data, headers=headers)
    if response.status_code == 200:
        return response.json().get('token')
    else:
        raise Exception(f"Erro na autenticação: {response.status_code} - {response.text}")

def consulta_completa(token, cpf_cnpj, data_base, authorized):
    """
    Realiza a consulta SCR completa.
    """
    query_url = f'{BASE_URL}/scr/complete/get/bc'
    params = {'cpf_cnpj': cpf_cnpj, 'data_base': data_base, 'authorized': authorized}
    headers = {'Authorization': f'Bearer {token}'}

    response = requests.get(query_url, params=params, headers=headers)
    if response.status_code == 200:
        return response.json()
    else:
        raise Exception(f"Erro na consulta completa: {response.status_code}")

def obter_dados_html(url):
    """
    Obtém dados de uma página HTML e converte para DataFrame.
    """
    response = requests.get(url)
    if response.status_code == 200:
        soup = BeautifulSoup(response.content, 'html.parser')
        return pd.read_html(str(soup))
    else:
        raise Exception(f"Erro ao acessar {url}: {response.status_code}")

def exportar_para_excel(dados, save_path):
    """
    Exporta os dados coletados para um arquivo Excel.
    """
    workbook = Workbook()

    # Adicionar aba "Links" com os URLs
    links_sheet = workbook.active
    links_sheet.title = "Links"
    links_sheet.append(["Origem", "URL"])
    
    for key, url in dados.items():
        # Adicionar link na aba "Links"
        links_sheet.append([key, url])

        try:
            html_tables = obter_dados_html(url)
            combined_df = pd.concat(html_tables, ignore_index=True)  # Combinar todas as tabelas de um link

            # Criar uma aba para o link e adicionar os dados
            sheet = workbook.create_sheet(title=key)
            for r_idx, row in enumerate(dataframe_to_rows(combined_df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    sheet.cell(row=r_idx, column=c_idx, value=value)
        except Exception as e:
            logging.error(f"Erro ao processar {key}: {e}")

    if 'Sheet' in workbook.sheetnames:
        del workbook['Sheet']
    workbook.save(save_path)
    logging.info(f"Arquivo salvo em: {save_path}")
